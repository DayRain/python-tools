#!/usr/bin/python
# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, colors, Alignment, Border, Side
import os


class ExcelData:
    def __init__(self, headers: list, data_list: list):
        self.headers = headers
        self.data_list = data_list


class ExcelConfig:
    def __init__(self, file_path, sheet_name, cover, header_font, header_bold, body_font, header_horizontal: bool,
                 body_horizontal: bool, border: bool):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.cover = cover
        self.header_font = header_font
        self.header_bold = header_bold
        self.body_font = body_font
        self.header_center = header_horizontal
        self.body_center = body_horizontal
        self.border = border


class ExcelModel:
    def __init__(self, config: ExcelConfig, excel_data: ExcelData):
        self.config = config
        self.excel_data = excel_data


def write(excel_model: ExcelModel):
    config = excel_model.config
    if os.path.exists(config.file_path) and not config.cover:
        wb = load_workbook(config.file_path)
        if config.sheet_name in wb.sheetnames:
            del wb[config.sheet_name]
        ws = wb.create_sheet(config.sheet_name, index=len(wb.sheetnames) + 1)
    else:
        wb = Workbook()
        ws = wb.active

    ws.title = config.sheet_name
    data = excel_model.excel_data

    ws.append(data.headers)

    for cells in data.data_list:
        ws.append(cells)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for irow, row in enumerate(ws.rows, start=1):
        if irow == 1:
            # header
            font = Font('黑体', bold=config.header_bold, sz=config.header_font)
            for cell in row:
                cell.font = font
                if config.header_center:
                    cell.alignment = Alignment(horizontal='center')

                if config.border:
                    cell.border = thin_border

        else:
            # body
            for cell in row:
                if config.body_center:
                    cell.alignment = Alignment(horizontal='center')
                font = Font(sz=config.body_font)
                cell.font = font

                if config.border:
                    cell.border = thin_border

    wb.save(config.file_path)


def read(file_name):
    wb = load_workbook(filename=file_name)
    ws = wb.active
    row_list = []
    for row in ws.rows:
        col_list = []
        for cell in row:
            col_list.append(cell.value)
        row_list.append(col_list)
    return row_list
